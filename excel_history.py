"""
excel_history.py
----------------

Updates:

- Global Audit History
- Department Audit History

This module NEVER creates a new history file.
It updates the existing ones.
"""


from __future__ import annotations
import pandas as pd

from pathlib import Path

import calendar

from openpyxl import load_workbook

from config import (
    GLOBAL_HISTORY_FILE,
    DEPARTMENT_HISTORY
)

from excel_formatter import (
    format_month_sheet
)
# ==========================================================
# MONTH
# ==========================================================

def get_month_name(audit_date):

    return calendar.month_name[
        audit_date.month
    ]
# ==========================================================
# FIND MONTH ROW
# ==========================================================

def find_month_row(
    worksheet,
    month
):
    """
    Finds the row where the month is written.
    """

    for row in range(1, worksheet.max_row + 1):

        value = worksheet.cell(
            row=row,
            column=2
        ).value

        if value is None:
            continue

        if str(value).strip().lower() == month.lower():

            return row

    raise ValueError(

        f"Month '{month}' not found."

    )

# ==========================================================
# KPI COLUMNS
# ==========================================================

KPI_COLUMNS = {

    "Internal Projects": 3,

    "External Projects": 4,

    "Projects with PO Consumed": 5,

    "Projects to Close This Year": 6,

    "Projects Without Hours Last 3 Months": 7,

    "Number of Samples": 8,

    "Pending PO Amount": 9,

    "Projects Without PM Hours": 10,

    "Delayed Projects": 11,

    "Samples In Delayed Projects": 12,

    "% Hours Reliability": 13,

    "% Users Excluded": 14

}
# ==========================================================
# WRITE KPI ROW
# ==========================================================

def write_kpis(

    worksheet,

    row,

    kpis,

    reliability,

    excluded

):

    for kpi, column in KPI_COLUMNS.items():

        if kpi == "% Hours Reliability":

            worksheet.cell(

                row=row,

                column=column

            ).value = reliability

            continue

        if kpi == "% Users Excluded":

            worksheet.cell(

                row=row,

                column=column

            ).value = excluded

            continue

        worksheet.cell(

            row=row,

            column=column

        ).value = kpis[kpi]
# ==========================================================
# MONTH SHEET
# ==========================================================

def create_month_sheet(

    workbook,

    sheet_name,

    dataframe

):

    if sheet_name in workbook.sheetnames:

        del workbook[sheet_name]

    worksheet = workbook.create_sheet(

        sheet_name

    )

    format_month_sheet(

        worksheet,

        dataframe

    )
# ==========================================================
# UPDATE HISTORY
# ==========================================================

def update_history(

    file,

    sheet_name,

    kpis,

    dataframe,

    audit_date,

    reliability,

    excluded

):

    workbook = load_workbook(file)

    worksheet = workbook[sheet_name]

    month = get_month_name(

        audit_date

    )

    row = find_month_row(

        worksheet,

        month

    )

    write_kpis(

        worksheet,

        row,

        kpis,

        reliability,

        excluded

    )

    create_month_sheet(

        workbook,

        audit_date.strftime("%B %Y"),

        dataframe

    )

    workbook.save(file)
# ==========================================================
# GLOBAL HISTORY
# ==========================================================

def update_global_history(
    audit,
    reliability,
    excluded,
    audit_date
):
    # -----------------------------------------
    # Update KPIs by department
    # -----------------------------------------

    for _, row in audit["summary"].iterrows():

        department = row["Department"]

        workbook = load_workbook(
            GLOBAL_HISTORY_FILE
        )

        worksheet = workbook[department]

        month = get_month_name(
            audit_date
        )

        row_number = find_month_row(
            worksheet,
            month
        )

        write_kpis(
            worksheet,
            row_number,
            row.to_dict(),
            reliability[department],
            excluded[department]
        )

        workbook.save(
            GLOBAL_HISTORY_FILE
        )

    # -----------------------------------------
    # Create global monthly sheet
    # -----------------------------------------

    global_dataframe = pd.concat(
        [
            audit["departments"][department]["all"]
            for department in audit["departments"]
        ],
        ignore_index=True
    )

    workbook = load_workbook(
        GLOBAL_HISTORY_FILE
    )

    create_month_sheet(
        workbook,
        audit_date.strftime("%B %Y"),
        global_dataframe
    )

    workbook.save(
        GLOBAL_HISTORY_FILE
    )
# ==========================================================
# DEPARTMENT HISTORY
# ==========================================================

def update_department_history(

    audit,

    reliability,

    excluded,

    audit_date

):

    for _, row in audit["summary"].iterrows():

        department = row["Department"]

        dataframe = audit["departments"][

            department

        ]["all"]

        update_history(

            DEPARTMENT_HISTORY[

                department

            ],

            department,

            row.to_dict(),

            dataframe,

            audit_date,

            reliability[department],

            excluded[department]

        )
# ==========================================================
# SAVE HISTORIES
# ==========================================================

def save_histories(

    audit,

    reliability,

    excluded,

    audit_date

):

    update_global_history(

        audit,

        reliability,

        excluded,

        audit_date

    )

    update_department_history(

        audit,

        reliability,

        excluded,

        audit_date

    )
