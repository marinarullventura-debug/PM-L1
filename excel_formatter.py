"""
excel_formatter.py
------------------

Functions for formatting Excel worksheets.

Responsibilities
----------------
- Write DataFrames
- Colour audit cells
- Auto-adjust column widths
- Freeze panes
- Apply filters
"""

from __future__ import annotations

from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment

from openpyxl.utils import get_column_letter
# ==========================================================
# STYLES
# ==========================================================

RED_FILL = PatternFill(
    fill_type="solid",
    start_color="FF9999",
    end_color="FF9999"
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    start_color="D9EAD3",
    end_color="D9EAD3"
)

HEADER_FONT = Font(
    bold=True
)

CENTER = Alignment(
    horizontal="center",
    vertical="center"
)
# ==========================================================
# WRITE DATAFRAME
# ==========================================================

def write_dataframe(
    worksheet,
    dataframe
):
    """
    Writes a dataframe into an openpyxl worksheet.
    """

    # Headers

    for col, column in enumerate(dataframe.columns, start=1):

        cell = worksheet.cell(
            row=1,
            column=col
        )

        cell.value = column
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    # Data

    for row_idx, (_, row) in enumerate(
        dataframe.iterrows(),
        start=2
    ):

        for col_idx, value in enumerate(
            row,
            start=1
        ):

            if isinstance(value, list):
                value = "; ".join(str(v) for v in value)

            elif not isinstance(
                value,
                (str, int, float, bool, type(None))
            ):
                value = str(value)

            worksheet.cell(
                row=row_idx,
                column=col_idx
            ).value = value

# ==========================================================
# PAINT INCIDENTS
# ==========================================================

def colour_audit_cells(
    worksheet,
    dataframe
):
    """
    Colours only the cells affected
    by the audit.
    """

    # Column name -> Excel column

    excel_columns = {

        column: idx + 1

        for idx, column

        in enumerate(dataframe.columns)

    }

    for row_idx, (_, row) in enumerate(

        dataframe.iterrows(),

        start=2

    ):

        for column in row["AUDIT_COLUMNS"]:

            if column not in excel_columns:

                continue

            excel_column = excel_columns[column]

            worksheet.cell(

                row=row_idx,

                column=excel_column

            ).fill = RED_FILL
# ==========================================================
# AUTO WIDTH
# ==========================================================

def auto_width(
    worksheet
):

    for column_cells in worksheet.columns:

        length = max(

            len(str(cell.value))

            if cell.value is not None

            else 0

            for cell in column_cells

        )

        worksheet.column_dimensions[

            get_column_letter(

                column_cells[0].column

            )

        ].width = length + 3
# ==========================================================
# FILTER
# ==========================================================

def add_filter(
    worksheet
):

    worksheet.auto_filter.ref = worksheet.dimensions
# ==========================================================
# FREEZE
# ==========================================================

def freeze(
    worksheet
):

    worksheet.freeze_panes = "A2"
# ==========================================================
# FORMAT SHEET
# ==========================================================

def format_month_sheet(

    worksheet,

    dataframe

):

    write_dataframe(

        worksheet,

        dataframe

    )

    colour_audit_cells(

        worksheet,

        dataframe

    )

    add_filter(

        worksheet

    )

    freeze(

        worksheet

    )

    auto_width(

        worksheet
    )